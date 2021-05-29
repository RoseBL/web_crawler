# 导入库
from openpyxl import Workbook, load_workbook

# 实例化
wb = Workbook()
# 获取当前active的sheet
sheet = wb.active
print(sheet.title)
sheet.title = "123"
print(sheet.title)
sheet["B9"] = "123"
sheet["C9"] = 170
sheet["A9"] = 110
sheet.append(["hello", "170", "180"])

wb.save("database/excel_test.xlsx")
# 打开已有文件
# wb = load_workbook("database/excel_test.xlsx")
